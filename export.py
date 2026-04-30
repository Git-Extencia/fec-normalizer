"""
Export FEC enrichi vers différents formats.

Stratégie par défaut (Excel-first, conforme aux usages du pôle Audit) :
- Toujours Excel si le volume est compatible avec la limite Excel (1 048 576 lignes)
- Avertissement utilisateur entre 500 000 et 1 000 000 lignes
- Au-dessus de 1 000 000 lignes : découpage automatique du classeur en plusieurs
  feuilles, par Exercice si la colonne existe, sinon par tranches.

Parquet reste accessible via format_force="parquet" pour l'Infocentre (phase 2)
ou pour les utilisateurs qui le préfèrent.
"""

from __future__ import annotations

import polars as pl
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# Au-delà de ce seuil, on prévient l'utilisateur que l'export va prendre
# du temps et que l'ouverture dans Excel sera lente.
SEUIL_AVERTISSEMENT_EXCEL = 500_000

# Limite physique d'Excel : 1 048 576 lignes par feuille. On garde une marge
# de sécurité en plafonnant à 1 000 000. Au-delà, on découpe en plusieurs feuilles.
LIMITE_EXCEL_MONOFEUILLE = 1_000_000

# Taille des tranches quand on découpe par paquets (cas où il n'y a pas de
# colonne Exercice utilisable pour un découpage métier).
TAILLE_TRANCHE = 900_000

# Ancien nom conservé pour rétrocompatibilité avec les imports existants.
SEUIL_EXCEL = LIMITE_EXCEL_MONOFEUILLE


def exporter(
    df: pl.DataFrame,
    chemin_sortie: str | Path,
    format_force: str | None = None,
) -> Path:
    """
    Exporte un FEC enrichi.

    Args:
        df: DataFrame enrichi
        chemin_sortie: Chemin de sortie (extension auto selon le format)
        format_force: 'xlsx', 'parquet', 'csv' ou None.
            None = défaut Excel (avec découpage si nécessaire).
            'xlsx' = force Excel quel que soit le volume (refus si > 1M et pas
                     de colonne Exercice exploitable).
            'parquet' = sortie Parquet compressée (utile Infocentre).
            'csv' = sortie CSV séparée par point-virgule.

    Returns:
        Chemin du fichier créé.
    """
    chemin = Path(chemin_sortie)

    # Excel devient le défaut, conformément à la décision du pôle Audit
    if format_force is None:
        format_force = "xlsx"

    if format_force == "xlsx":
        chemin = chemin.with_suffix(".xlsx")
        _exporter_excel(df, chemin)
    elif format_force == "parquet":
        chemin = chemin.with_suffix(".parquet")
        df.write_parquet(chemin, compression="snappy")
    elif format_force == "csv":
        chemin = chemin.with_suffix(".csv")
        df.write_csv(chemin, separator=";")
    else:
        raise ValueError(f"Format inconnu : {format_force}")

    return chemin


def _exporter_excel(df: pl.DataFrame, chemin: Path) -> None:
    """
    Export Excel formaté (charte Extencia).

    Comportement adaptatif au volume :
    - <= 1 000 000 lignes : monofeuille classique
    - > 1 000 000 lignes : découpage automatique
        - en feuilles par Exercice si la colonne existe
        - sinon en tranches de 900 000 lignes
    """
    if df.height <= LIMITE_EXCEL_MONOFEUILLE:
        if df.height >= SEUIL_AVERTISSEMENT_EXCEL:
            print(
                f"⚠️  Volume élevé ({df.height:,} lignes) — l'export Excel "
                f"peut prendre 1 à 3 minutes et l'ouverture dans Excel sera lente."
                .replace(",", " ")
            )
        wb = Workbook()
        ws = wb.active
        ws.title = "FEC_enrichi"
        _ecrire_feuille_excel(ws, df)
    else:
        # Découpage obligatoire : on dépasse la limite physique d'Excel
        wb = Workbook()
        wb.remove(wb.active)  # On supprime la feuille par défaut
        partitions = _decouper_pour_excel(df)
        nb_feuilles = len(partitions)
        print(
            f"⚠️  Volume très élevé ({df.height:,} lignes) — découpage automatique "
            f"en {nb_feuilles} feuilles pour respecter la limite Excel."
            .replace(",", " ")
        )
        for nom_feuille, df_partition in partitions:
            ws = wb.create_sheet(title=nom_feuille[:31])  # Excel limite à 31 caractères
            _ecrire_feuille_excel(ws, df_partition)

    wb.save(chemin)


def _decouper_pour_excel(df: pl.DataFrame) -> list[tuple[str, pl.DataFrame]]:
    """
    Découpe un DataFrame trop gros pour Excel monofeuille en partitions.

    Stratégie :
    - Si la colonne Exercice existe et permet un découpage propre, on la suit.
    - Sinon on découpe par tranches de TAILLE_TRANCHE lignes (numérotées).

    Returns:
        Liste de tuples (nom_feuille, dataframe_partition)
    """
    # Stratégie 1 : découpage métier par Exercice
    if "Exercice" in df.columns:
        exercices = sorted(df["Exercice"].unique().to_list())
        partitions = []
        for ex in exercices:
            sous_df = df.filter(pl.col("Exercice") == ex)
            if sous_df.height <= LIMITE_EXCEL_MONOFEUILLE:
                partitions.append((f"Exercice_{ex}", sous_df))
            else:
                # Un exercice qui dépasse 1M lignes : on retombe sur le tranchage
                partitions.extend(
                    _decouper_par_tranches(sous_df, prefix=f"Exercice_{ex}")
                )
        return partitions

    # Stratégie 2 : découpage par tranches numérotées
    return _decouper_par_tranches(df)


def _decouper_par_tranches(
    df: pl.DataFrame,
    prefix: str = "Tranche",
) -> list[tuple[str, pl.DataFrame]]:
    """Découpe un DataFrame en tranches de TAILLE_TRANCHE lignes."""
    partitions = []
    for i in range(0, df.height, TAILLE_TRANCHE):
        sous_df = df.slice(i, TAILLE_TRANCHE)
        numero = i // TAILLE_TRANCHE + 1
        partitions.append((f"{prefix}_{numero}", sous_df))
    return partitions


def _ecrire_feuille_excel(ws, df: pl.DataFrame) -> None:
    """
    Écrit un DataFrame dans une feuille Excel avec mise en forme charte Extencia
    (en-tête bleu nuit, filtres auto, freeze panes, largeurs adaptatives).
    """
    colonnes = df.columns

    # En-tête
    for col_idx, nom_col in enumerate(colonnes, start=1):
        cell = ws.cell(row=1, column=col_idx, value=nom_col)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", start_color="152639")  # Bleu Extencia
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Données
    for ligne in df.iter_rows():
        ws.append(list(ligne))

    # Filtres auto + freeze première ligne
    if df.height > 0:
        ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # Largeurs adaptatives (sample 100 lignes pour rester rapide)
    sample = df.head(100)
    for col_idx, nom_col in enumerate(colonnes, start=1):
        max_len = max(
            len(str(nom_col)),
            sample[nom_col].cast(pl.Utf8).str.len_chars().max() or 10,
        )
        largeur = min(max_len + 2, 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = largeur


def exporter_rapport_diagnostic(
    diagnostics: list[dict],
    chemin_sortie: str | Path,
) -> Path:
    """
    Exporte un rapport de diagnostic pour traçabilité audit.
    Liste tous les FEC traités avec leurs caractéristiques.
    """
    chemin = Path(chemin_sortie).with_suffix(".xlsx")
    df_diag = pl.DataFrame(diagnostics)

    wb = Workbook()
    ws = wb.active
    ws.title = "Diagnostic_FEC"

    for col_idx, nom_col in enumerate(df_diag.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=nom_col)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="5EB2A1")  # Vert Extencia

    for ligne in df_diag.iter_rows():
        ws.append([str(v) if isinstance(v, list) else v for v in ligne])

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    for col_idx in range(1, df_diag.width + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 25

    wb.save(chemin)
    return chemin
